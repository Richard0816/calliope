"""calliope.tabs.crosscorrelation.tab - Tab 7: cluster x cluster
cross-correlation.

What this tab is for
--------------------
Tab 6 told us *which* cells co-fire (ensembles). Tab 7 tells us
*who leads whom*: for each pair of clusters Cᵢ x Cⱼ, and for each
pair of cells (one from each cluster), we shift cell A's trace by
every lag from -L..+L frames, take the Pearson r against B at each
shift, and record:

* **best_lag_sec**: the lag at which correlation peaks (positive ->
  A leads B).
* **max_corr**: that peak Pearson r.
* **zero_lag_corr**: Pearson r at lag 0 (instantaneous co-firing).

Two modes
---------
* **Full recording** -- one cross-correlation pass over every pair
  using the entire dF/F trace.
* **Per event** -- repeat the analysis cropped to each event window
  detected by Tab 5 (so we can ask "during *this* seizure, who led
  whom?"). Event windows arrive live via
  ``AppState.event_results`` (Tab 5 publishes after every render);
  the on-disk EventWindows xlsx sheet acts as a cold-start fallback
  when no Tab 5 publish exists yet.

Both modes use the batched implementation in
``calliope.core.crosscorrelation.batch_xcorr_clusters``:
*one matmul per lag* covers every ROI pair in a cluster pair
simultaneously. So a 5x10 cluster pair x 100 lags is 100 matmuls,
not 5000 individual cross-correlations.

What the user sees
------------------
- Path / prefix / cluster folder / fps entries.
- Search parameters (max lag, zero-lag toggle, GPU toggle).
- Two big "Run" buttons -- one for each mode -- plus an Abort.
- A progress bar driven by the worker's progress callback.
- A live console at the bottom showing per-batch timing.
- Two extra buttons:
    * **Show violins** opens a separate window summarising the
      per-pair best-lag and zero-lag distributions across cluster
      pairs as violin plots, with a sign-flip permutation test for
      "is this lead-lag significantly nonzero".
    * **Single-pair preview** plots the cross-correlation curve for
      one ROI pair you pick from a dropdown -- handy for sanity
      checks.

Outputs
-------
For each (Cᵢ x Cⱼ) the tab writes one CSV per pair under
``<prefix>cluster_results/<cluster_folder>/cross_correlation_full/``.
Per-event mode writes parallel folders per event window. Each row
has best_lag_sec, max_corr, zero_lag_corr for one ROI pair.

Threading
---------
Heavy lifting runs on a worker thread; the worker calls a progress
callback that polls a ``threading.Event`` set by the Abort button so
long runs can be cancelled cleanly via the ``RunAborted`` exception
defined below.
"""

from __future__ import annotations

import queue
import threading
import tkinter as tk
import traceback
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Optional

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk,
)

import customtkinter as ctk

from ...gui_common import (
    drain_queue, install_scroll_router, restyle_matplotlib_toolbar,
)
from .logic import xc
from .logic import utils


DEFAULT_PREFIX = "r0p7_filtered_"
# Cluster ROI .npy files now live directly in
# ``<plane0>/<prefix>cluster_results/`` (no ``gui_recluster`` subdir).
# Empty cluster_folder is interpreted by xc.run_cluster_xcorr_*_fast
# as "no extra subdir" -- the path join is conditional on truthiness.
DEFAULT_CLUSTER_FOLDER = ""
DEFAULT_MAX_LAG_S = 2.0
DEFAULT_FULL_OUTPUT_SUBDIR = "cross_correlation_full"
POLL_MS = 100


class RunAborted(Exception):
    """Raised inside the progress callback when the user clicks Abort.

    Propagates out of the cross-correlation routine, gets caught by the
    worker's outer try/except, and surfaces as an ``aborted`` queue
    message instead of an error dialog.

    Implementation note
    -------------------
    The cross-correlation routines in ``core.crosscorrelation`` accept
    an optional ``progress_cb(done, total)`` callback they invoke
    after every batch. The Abort button sets a
    ``threading.Event``; the callback checks it and raises this
    exception when the flag is set, unwinding out of the heavy loop
    cleanly. Subclassing ``Exception`` (instead of ``BaseException``)
    means it gets caught by ordinary ``except Exception:`` blocks --
    important so we don't accidentally trigger a "hard" abort that
    propagates past the worker thread's outer ``except``.
    """
    pass


# ---------------------------------------------------------------------------
# Violin-plot helpers (mirrors fig6_real_paper.plot_violin so the standalone
# paper figure script and this tab stay visually consistent).
# ---------------------------------------------------------------------------


def _natural_pair_key(name: str) -> tuple[int, int]:
    """Sort 'C1xC10' before 'C2xC3' by parsing the trailing integers."""
    a, b = name.split("x")
    return int(a[1:]), int(b[1:])


def _load_pair_data(xcorr_root: Path) -> dict[str, pd.DataFrame]:
    pair_data: dict[str, pd.DataFrame] = {}
    for d in sorted(xcorr_root.iterdir()):
        if not d.is_dir():
            continue
        csvs = list(d.glob("*_summary.csv"))
        if not csvs:
            continue
        try:
            pair_data[d.name] = pd.read_csv(csvs[0])
        except Exception:
            continue
    try:
        return dict(sorted(pair_data.items(),
                           key=lambda kv: _natural_pair_key(kv[0])))
    except Exception:
        return pair_data


def _get_metric_arrays(pair_data: dict[str, pd.DataFrame], metric: str):
    labels: list[str] = []
    arrays: list[np.ndarray] = []
    for pair, df in pair_data.items():
        if metric not in df.columns:
            continue
        arr = df[metric].dropna().to_numpy()
        if arr.size == 0:
            continue
        labels.append(pair)
        arrays.append(arr)
    return labels, arrays


def _sign_flip_pvalue(values: np.ndarray, n_perm: int = 10000,
                      seed: int = 0, chunk: int = 256) -> float:
    # Chunked to keep peak memory at ~chunk*N instead of n_perm*N.
    # The naive (n_perm, N) allocation OOMs the GUI when N is large
    # (e.g. 200x200 ROI cluster pairs -> N=40_000, signs alone = 3.2 GB).
    values = np.asarray(values, dtype=np.float64)
    n = values.size
    if n == 0:
        return 1.0
    rng = np.random.default_rng(seed)
    observed = abs(float(values.mean()))
    inv_n = 1.0 / n
    hits = 0
    done = 0
    while done < n_perm:
        m = min(chunk, n_perm - done)
        signs = rng.integers(0, 2, size=(m, n), dtype=np.int8)
        signs <<= 1            # 0,1 -> 0,2
        signs -= 1             # 0,2 -> -1,+1
        permuted = signs.astype(np.float64) @ values
        permuted *= inv_n
        hits += int(np.count_nonzero(np.abs(permuted) >= observed))
        done += m
    return hits / n_perm


def _add_significance(ax, positions, arrays) -> None:
    pvals = [_sign_flip_pvalue(arr) for arr in arrays]
    ymax = ax.get_ylim()[1]
    for i, p in enumerate(pvals):
        if p < 0.05:
            ax.text(positions[i], ymax * 0.93, "*",
                    ha="center", va="bottom", fontsize=14)


def _significance_marker(p: float) -> str:
    if p < 0.001:
        return "***"
    if p < 0.01:
        return "**"
    if p < 0.05:
        return "*"
    return ""


# Lead/Lag/NS palette (matches the reference paper figure).
_LEAD_COLOR = "#6FA8FF"   # blue   - mean lag sig. > 0  (ROI A leads)
_LAG_COLOR = "#E87B73"    # red    - mean lag sig. < 0  (ROI A lags)
_NS_COLOR = "#CFCFCF"     # gray   - not significant


def _plot_violin(ax, labels, arrays, ylabel: str, title: str,
                 show_sig: bool = False) -> None:
    pos = np.arange(1, len(arrays) + 1)
    ax.violinplot(arrays, positions=pos,
                  showmedians=True, showextrema=False)
    ax.set_xticks(pos)
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    ax.axhline(0, linestyle="--", linewidth=1)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    if show_sig:
        _add_significance(ax, pos, arrays)


def _plot_lag_violin(ax, labels, arrays, ylabel: str, title: str) -> None:
    """Lead/Lag-styled violin plot for per-pair best-lag distributions.

    Each violin is coloured by the sign of its mean and the sign-flip
    permutation p-value:
      * blue  - mean > 0 and p < 0.05 (Lead)
      * red   - mean < 0 and p < 0.05 (Lag)
      * gray  - not significant
    Significance markers (*, **, ***) sit above each violin in matching colour.
    A short black tick marks the median inside each violin.
    """
    from matplotlib.patches import Patch
    if not arrays:
        return
    pos = np.arange(1, len(arrays) + 1)
    pvals = [_sign_flip_pvalue(arr) for arr in arrays]
    means = [float(np.mean(arr)) for arr in arrays]

    colors: list[str] = []
    for p, m in zip(pvals, means):
        if p < 0.05 and m > 0:
            colors.append(_LEAD_COLOR)
        elif p < 0.05 and m < 0:
            colors.append(_LAG_COLOR)
        else:
            colors.append(_NS_COLOR)

    parts = ax.violinplot(arrays, positions=pos,
                          showmedians=False, showextrema=False,
                          widths=0.85)
    for body, color in zip(parts["bodies"], colors):
        body.set_facecolor(color)
        body.set_edgecolor("black")
        body.set_alpha(0.9)
        body.set_linewidth(0.6)

    # Median tick (short horizontal black bar) inside each violin.
    for i, arr in enumerate(arrays):
        med = float(np.median(arr))
        ax.hlines(med, pos[i] - 0.28, pos[i] + 0.28,
                  colors="black", linewidth=1.6, zorder=4)

    # Tighten y-limits and place significance markers above the data.
    all_vals = np.concatenate(arrays)
    vmin = float(np.min(all_vals))
    vmax = float(np.max(all_vals))
    span = max(vmax - vmin, 1e-9)
    y_top = vmax + span * 0.18
    y_bot = vmin - span * 0.05
    ax.set_ylim(y_bot, y_top)
    marker_y = vmax + span * 0.08
    for i, p in enumerate(pvals):
        marker = _significance_marker(p)
        if not marker:
            continue
        ax.text(pos[i], marker_y, marker,
                ha="center", va="center", fontsize=12,
                color=colors[i], fontweight="bold")

    ax.set_xticks(pos)
    ax.set_xticklabels(labels, rotation=30, ha="right")
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    ax.axhline(0, linestyle="--", linewidth=1, color="black")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    legend_handles = [
        Patch(facecolor=_LEAD_COLOR, edgecolor="black", label="Lead (sig.)"),
        Patch(facecolor=_LAG_COLOR, edgecolor="black", label="Lag (sig.)"),
        Patch(facecolor=_NS_COLOR, edgecolor="black", label="n.s."),
    ]
    ax.legend(handles=legend_handles, loc="upper right",
              frameon=False, ncol=3, fontsize=8)


class ViolinWindow(ctk.CTkToplevel):
    """Inspection window that renders zero-lag-correlation and best-lag
    violins per cluster pair from full-recording cross-correlation outputs.

    The "Include clusters" dropdown lists every cluster id present in the
    pair-summary CSV file names. Unticking a cluster drops every pair that
    involves it (e.g. unticking C3 removes C1xC3, C2xC3, C3xC3, C3xC4, ...);
    the figure re-renders on every toggle."""

    def __init__(self, master, *, xcorr_root: Path,
                 recording_label: str = "") -> None:
        super().__init__(master)
        self.title(f"Cross-correlation violins  -  {xcorr_root.name}"
                   + (f"  ({recording_label})" if recording_label else ""))
        self.geometry("1000x740")
        self.wm_minsize(720, 540)

        self._xcorr_root = xcorr_root
        # Recording label = human-readable name shown in the title
        # bar. ``xcorr_root`` lives at
        # ``<plane0>/<prefix>cluster_results/<cluster_folder>/cross_correlation_full``
        # so two ``.parent`` walks land on plane0 -- and from plane0
        # the recording id is no longer always two more parents up
        # (the sparse_plus_cellpose layout adds ``detection/final``
        # in between). Use the robust resolver instead.
        if recording_label:
            self._recording_label = recording_label
        else:
            from ...core.utils import infer_recording_id
            self._recording_label = infer_recording_id(xcorr_root)
        self._pair_data = _load_pair_data(xcorr_root)

        if not self._pair_data:
            # Theme-safe error red -- "red" on the dark theme reads too
            # dim; this hex pops against the gray17 surround.
            ctk.CTkLabel(
                self, text_color="#ff6b6b",
                text=("No CAxCB/*_summary.csv files found in:\n"
                      f"{xcorr_root}\n\n"
                      "Run the full-recording cross-correlation first."),
                wraplength=900,
            ).pack(fill="both", expand=True, padx=20, pady=20)
            return

        # Cluster ids present in the pair names (e.g. {1, 2, 3, 4, 5}).
        cluster_ids: set[int] = set()
        for pair in self._pair_data:
            try:
                a, b = _natural_pair_key(pair)
                cluster_ids.update((a, b))
            except Exception:
                continue
        self._cluster_ids = sorted(cluster_ids)
        self._cluster_vars: dict[int, tk.BooleanVar] = {}
        self._applied_clusters: frozenset[int] = frozenset(self._cluster_ids)

        self._build_ui()
        self._render()
        # Consume wheel events so they don't leak through CTk's
        # global bind_all and scroll the main app behind this popout.
        install_scroll_router(self)

    def _build_ui(self) -> None:
        ctl = ctk.CTkFrame(self, fg_color="transparent")
        ctl.pack(fill="x", padx=6, pady=(6, 0))
        ctk.CTkLabel(ctl, text="Include clusters:").pack(side="left")
        self.cluster_menu = tk.Menu(self, tearoff=False)
        # ttk.Menubutton because customtkinter has no cascade-menu widget --
        # we need the checkbutton-list dropdown for "include cluster" toggles.
        self.cluster_menu_btn = ttk.Menubutton(ctl, text="all", width=24)
        self.cluster_menu_btn.config(menu=self.cluster_menu)
        self.cluster_menu_btn.pack(side="left", padx=(8, 0))

        for cid in self._cluster_ids:
            var = tk.BooleanVar(value=True)
            self._cluster_vars[cid] = var
            self.cluster_menu.add_checkbutton(
                label=f"C{cid}", variable=var,
                onvalue=True, offvalue=False,
                command=self._on_cluster_toggle)
        if self._cluster_ids:
            self.cluster_menu.add_separator()
            self.cluster_menu.add_command(
                label="Select all", command=self._select_all_clusters)
            self.cluster_menu.add_command(
                label="Clear", command=self._clear_clusters)

        self.apply_btn = ctk.CTkButton(
            ctl, text="Apply", width=90,
            command=self._on_apply_clusters, state="disabled")
        self.apply_btn.pack(side="left", padx=(6, 0))

        self._pair_count_var = tk.StringVar(value="")
        ctk.CTkLabel(ctl, textvariable=self._pair_count_var
                     ).pack(side="left", padx=(12, 0))

        self.fig = plt.Figure(figsize=(10, 7), constrained_layout=True)
        self.ax_corr = self.fig.add_subplot(2, 1, 1)
        self.ax_lag = self.fig.add_subplot(2, 1, 2)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self)

        tb_frame = ctk.CTkFrame(self, fg_color="transparent")
        tb_frame.pack(fill="x")
        self.toolbar = NavigationToolbar2Tk(
            self.canvas, tb_frame, pack_toolbar=False)
        self.toolbar.update()
        restyle_matplotlib_toolbar(self.toolbar)
        self.toolbar.pack(side="left", fill="x")
        ctk.CTkButton(
            tb_frame, text="Save data...", width=110,
            command=self._on_save_violin_data,
        ).pack(side="left", padx=(8, 0))
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

    # -- Data export -------------------------------------------------------

    def _on_save_violin_data(self) -> None:
        """Dump the currently-filtered per-pair lag and zero-lag-corr
        distributions to a single wide CSV. Violins are PolyCollection
        objects so the generic ``save_figure_data`` helper can't recover
        them; this writes the raw arrays directly."""
        from ... import plot_data_export
        pair_data = self._filter_pair_data()
        corr_labels, corr_arrays = _get_metric_arrays(
            pair_data, "zero_lag_corr")
        lag_labels, lag_arrays = _get_metric_arrays(
            pair_data, "best_lag_sec")
        cols: dict[str, np.ndarray] = {}
        for label, arr in zip(corr_labels, corr_arrays):
            cols[f"{label}__zero_lag_corr"] = np.asarray(arr, dtype=float)
        for label, arr in zip(lag_labels, lag_arrays):
            cols[f"{label}__best_lag_sec"] = np.asarray(arr, dtype=float)
        plot_data_export.save_columns_csv(
            self, cols,
            default_basename=f"violin_data_{self._xcorr_root.name}")

    # -- Cluster picker ----------------------------------------------------

    def _included_clusters(self) -> set[int]:
        return {cid for cid, v in self._cluster_vars.items() if v.get()}

    def _filter_pair_data(self) -> dict[str, pd.DataFrame]:
        included = self._included_clusters()
        out: dict[str, pd.DataFrame] = {}
        for pair, df in self._pair_data.items():
            try:
                a, b = _natural_pair_key(pair)
            except Exception:
                continue
            if a in included and b in included:
                out[pair] = df
        return out

    def _update_button_label(self, n_pairs: int) -> None:
        included = sorted(self._included_clusters())
        total = len(self._cluster_ids)
        if not included:
            self.cluster_menu_btn.config(text="(none)")
        elif len(included) == total:
            self.cluster_menu_btn.config(text=f"all ({total})")
        elif len(included) <= 4:
            self.cluster_menu_btn.config(
                text=", ".join(f"C{c}" for c in included))
        else:
            self.cluster_menu_btn.config(
                text=f"{len(included)}/{total} clusters")
        self._pair_count_var.set(
            f"({n_pairs} pair{'' if n_pairs == 1 else 's'})")

    def _mark_pending(self) -> None:
        # Track whether the in-menu selection differs from what's plotted
        # so the Apply button only lights up when there's work to do.
        current = frozenset(self._included_clusters())
        if current != self._applied_clusters:
            self.apply_btn.configure(state="normal")
        else:
            self.apply_btn.configure(state="disabled")

    def _on_cluster_toggle(self) -> None:
        self._mark_pending()

    def _select_all_clusters(self) -> None:
        for v in self._cluster_vars.values():
            v.set(True)
        self._mark_pending()

    def _clear_clusters(self) -> None:
        for v in self._cluster_vars.values():
            v.set(False)
        self._mark_pending()

    def _on_apply_clusters(self) -> None:
        self._render()
        self.apply_btn.configure(state="disabled")

    # -- Render ------------------------------------------------------------

    def _render(self) -> None:
        self._applied_clusters = frozenset(self._included_clusters())
        pair_data = self._filter_pair_data()
        corr_labels, corr_arrays = _get_metric_arrays(
            pair_data, "zero_lag_corr")
        lag_labels, lag_arrays = _get_metric_arrays(
            pair_data, "best_lag_sec")

        self._update_button_label(len(pair_data))

        for ax in (self.ax_corr, self.ax_lag):
            ax.clear()
            ax.set_axis_on()

        title_suffix = self._recording_label
        if corr_arrays:
            _plot_violin(
                self.ax_corr, corr_labels, corr_arrays,
                "Zero-lag correlation",
                f"Per-pair zero-lag correlation  ({title_suffix})",
                show_sig=False,
            )
        else:
            self.ax_corr.set_axis_off()
            msg = ("No pairs match the current filter."
                   if not pair_data else "No 'zero_lag_corr' column")
            self.ax_corr.text(0.5, 0.5, msg,
                              ha="center", va="center",
                              transform=self.ax_corr.transAxes)

        if lag_arrays:
            _plot_lag_violin(
                self.ax_lag, lag_labels, lag_arrays,
                "Lag at peak correlation (s)",
                f"Per-pair best-lag distribution  ({title_suffix})",
            )
        else:
            self.ax_lag.set_axis_off()
            msg = ("No pairs match the current filter."
                   if not pair_data else "No 'best_lag_sec' column")
            self.ax_lag.text(0.5, 0.5, msg,
                             ha="center", va="center",
                             transform=self.ax_lag.transAxes)

        self.canvas.draw_idle()


def _read_event_windows_from_summary(plane0: Path):
    """Pull (start_s, end_s) tuples from the EventWindows sheet written by
    ``summary_writer.write_events_sheets``. Returns [] if the sheet or
    workbook is missing.
    """
    try:
        import openpyxl  # noqa
    except Exception:
        return []
    candidates = sorted(plane0.glob("*summary*.xlsx"))
    for path in candidates:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        if "EventWindows" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["EventWindows"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            continue
        header = [str(c).strip().lower() if c is not None else ""
                  for c in rows[0]]
        try:
            i_s = header.index("start_s")
            i_e = header.index("end_s")
        except ValueError:
            continue
        out = []
        for r in rows[1:]:
            try:
                out.append((float(r[i_s]), float(r[i_e])))
            except (TypeError, ValueError):
                continue
        if out:
            return out
    return []


class CrossCorrelationTab(ctk.CTkFrame):
    """Cluster x cluster cross-correlation, full recording + per-event."""

    def __init__(self, master, state=None) -> None:
        super().__init__(master, fg_color="transparent")
        self.state = state

        self._plane0: Optional[Path] = None
        self._prefix = DEFAULT_PREFIX
        self._cluster_folder = DEFAULT_CLUSTER_FOLDER
        self._fps: float = 15.07
        self._event_windows: list[tuple[float, float]] = []

        # Cache for the single-pair preview plot (avoid reopening memmap).
        self._dff_cache: Optional[np.memmap] = None
        self._dff_cache_key = None  # (plane0, prefix)

        self._q: queue.Queue = queue.Queue()
        self._worker: Optional[threading.Thread] = None
        # Abort flag: set by the Abort button, polled by the per-batch
        # progress_cb. Raises RunAborted to unwind the worker cleanly.
        self._abort_event: threading.Event = threading.Event()

        self._build_ui()
        drain_queue(self, self._q,
                    {"progress": self._on_xc_progress,
                     "done_full": self._on_xc_done_full,
                     "done_per_event": self._on_xc_done_per_event,
                     "aborted": self._on_xc_aborted,
                     "error": self._on_xc_error},
                    poll_ms=POLL_MS)

        if state is not None:
            try:
                state.subscribe_plane0(self._on_plane0_broadcast)
                state.subscribe_lowpass_ready(self._on_plane0_broadcast)
                state.subscribe_event_results(
                    self._on_event_results_broadcast)
                if getattr(state, "lowpass_plane0", None) is not None:
                    self._on_plane0_broadcast(state.lowpass_plane0)
                elif getattr(state, "plane0", None) is not None:
                    self._on_plane0_broadcast(state.plane0)
                if getattr(state, "event_results", None) is not None:
                    self._on_event_results_broadcast(state.event_results)
            except Exception:
                pass

    # -- UI ----------------------------------------------------------------

    def _build_ui(self) -> None:
        head = ctk.CTkFrame(self)
        head.pack(fill="x", padx=10, pady=(10, 6))
        ctk.CTkLabel(
            head,
            text="Cluster x cluster cross-correlation "
                 "(batched matmul, best lag + zero lag)",
            font=ctk.CTkFont(weight="bold")).pack(
            anchor="w", padx=8, pady=(6, 0))

        row1 = ctk.CTkFrame(head, fg_color="transparent")
        row1.pack(fill="x", padx=8, pady=2)
        ctk.CTkLabel(row1, text="plane0:").pack(side="left")
        self.path_var = tk.StringVar(value="")
        ctk.CTkEntry(row1, textvariable=self.path_var, width=560).pack(
            side="left", padx=(4, 4), fill="x", expand=True)
        ctk.CTkButton(row1, text="Browse...", width=90,
                      command=self._on_browse).pack(side="left")

        row2 = ctk.CTkFrame(head, fg_color="transparent")
        row2.pack(fill="x", padx=8, pady=2)
        ctk.CTkLabel(row2, text="prefix:").pack(side="left")
        self.prefix_var = tk.StringVar(value=DEFAULT_PREFIX)
        ctk.CTkEntry(row2, textvariable=self.prefix_var, width=140).pack(
            side="left", padx=(4, 12))

        ctk.CTkLabel(row2, text="cluster folder:").pack(side="left")
        self.cfolder_var = tk.StringVar(value=DEFAULT_CLUSTER_FOLDER)
        ctk.CTkEntry(row2, textvariable=self.cfolder_var, width=140).pack(
            side="left", padx=(4, 12))

        ctk.CTkLabel(row2, text="fps:").pack(side="left")
        self.fps_var = tk.StringVar(value="15.07")
        ctk.CTkEntry(row2, textvariable=self.fps_var, width=70).pack(
            side="left", padx=(4, 0))

        # Algorithm parameters.
        row3 = ctk.CTkFrame(head)
        row3.pack(fill="x", padx=8, pady=(6, 6))
        ctk.CTkLabel(row3, text="Search parameters",
                     font=ctk.CTkFont(weight="bold")).pack(
            anchor="w", padx=8, pady=(6, 0))

        row3_inner = ctk.CTkFrame(row3, fg_color="transparent")
        row3_inner.pack(fill="x", padx=8, pady=(0, 6))
        ctk.CTkLabel(row3_inner, text="max_lag (s):").pack(side="left")
        self.maxlag_var = tk.StringVar(value=str(DEFAULT_MAX_LAG_S))
        ctk.CTkEntry(row3_inner, textvariable=self.maxlag_var, width=70).pack(
            side="left", padx=(4, 12))

        self.zero_lag_var = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(row3_inner, text="also output zero-lag corr",
                        variable=self.zero_lag_var).pack(
            side="left", padx=(4, 12))

        self.gpu_var = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(row3_inner, text="use GPU if available",
                        variable=self.gpu_var).pack(side="left")

        # Circular-shift null for per-pair p-values. 500 is the audit-
        # recommended default for sub-1% p-value resolution; set to 0 to
        # disable (faster, no significance test). Adds an extra
        # ``p_value`` / ``p_value_fdr`` column to the summary CSV.
        # Reference: Cheng et al. eLife 2023, doi:10.7554/eLife.81279.
        ctk.CTkLabel(row3_inner, text="    shuffles (0=off):").pack(side="left")
        self.n_shuffles_var = tk.StringVar(value="500")
        ctk.CTkEntry(row3_inner, textvariable=self.n_shuffles_var,
                     width=70).pack(side="left", padx=(4, 0))

        # Partial correlation. When checked, the population-wide partial-
        # correlation matrix is computed once over the union of all
        # cluster ROIs (conditioning on every other ROI), then sliced
        # into a ``partial_corr_zero_lag`` column on each cluster pair's
        # summary CSV. Skipped automatically if N_union >= 0.5 * T
        # (precision matrix would be ill-conditioned). Reference: Sutera
        # et al. arXiv:1406.7865.
        self.partial_corr_var = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(
            row3_inner, text="+ partial corr (requires N < T/2)",
            variable=self.partial_corr_var,
        ).pack(side="left", padx=(12, 0))

        # Run buttons
        run_frame = ctk.CTkFrame(self)
        run_frame.pack(fill="x", padx=10, pady=(0, 6))
        ctk.CTkLabel(run_frame, text="Run",
                     font=ctk.CTkFont(weight="bold")).pack(
            anchor="w", padx=8, pady=(6, 0))
        run_inner = ctk.CTkFrame(run_frame, fg_color="transparent")
        run_inner.pack(fill="x", padx=8, pady=(0, 6))

        self.full_btn = ctk.CTkButton(
            run_inner, text="Run full-recording cross-correlation",
            command=self._on_run_full, state="disabled")
        self.full_btn.pack(side="left")

        self.per_event_btn = ctk.CTkButton(
            run_inner, text="Run per-event cross-correlation",
            command=self._on_run_per_event, state="disabled")
        self.per_event_btn.pack(side="left", padx=(8, 0))

        self.refresh_btn = ctk.CTkButton(
            run_inner, text="Reload event windows", width=160,
            command=self._on_refresh_events)
        self.refresh_btn.pack(side="left", padx=(8, 0))

        self.reload_btn = ctk.CTkButton(
            run_inner, text="Reload dF/F & clusters", width=170,
            command=self._on_reload_inputs)
        self.reload_btn.pack(side="left", padx=(8, 0))

        self.violin_btn = ctk.CTkButton(
            run_inner, text="Violin plot", width=100,
            command=self._on_violin, state="disabled")
        self.violin_btn.pack(side="left", padx=(8, 0))

        self.export_figs_btn = ctk.CTkButton(
            run_inner, text="Export figures", width=130,
            command=self._on_export_figures, state="disabled")
        self.export_figs_btn.pack(side="left", padx=(8, 0))

        # Abort button: only enabled while a worker is alive.
        self.abort_btn = ctk.CTkButton(
            run_inner, text="Abort", width=80,
            command=self._on_abort, state="disabled")
        self.abort_btn.pack(side="left", padx=(8, 0))

        self.event_count_var = tk.StringVar(value="events: -")
        ctk.CTkLabel(run_inner, textvariable=self.event_count_var).pack(
            side="left", padx=(12, 0))

        # Progress + status. CTkProgressBar is always on a 0..1 scale;
        # ``_set_progress`` below normalises ``done / total`` so callers
        # can keep thinking in "step N of M" terms.
        prog = ctk.CTkFrame(self, fg_color="transparent")
        prog.pack(fill="x", padx=10, pady=(0, 6))
        self.progress = ctk.CTkProgressBar(prog, mode="determinate", width=320)
        self.progress.set(0)
        self.progress.pack(side="left")
        self.status_var = tk.StringVar(value="Pick a plane0 folder.")
        ctk.CTkLabel(prog, textvariable=self.status_var,
                     anchor="w",
                     font=ctk.CTkFont(size=11, slant="italic")).pack(
            side="left", padx=(8, 0), fill="x", expand=True)

        # Bottom: log on the left, single-pair preview on the right.
        # ttk.PanedWindow because customtkinter has no resizable-pane
        # widget -- the drag handle between the two panels is the whole
        # reason this isn't a plain CTkFrame split.
        bottom = ttk.PanedWindow(self, orient="horizontal")
        bottom.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # -- Log pane
        log_frame = ctk.CTkFrame(bottom)
        bottom.add(log_frame, weight=1)
        ctk.CTkLabel(log_frame, text="Log",
                     font=ctk.CTkFont(weight="bold")).pack(
            anchor="w", padx=8, pady=(6, 0))
        self.log_text = ctk.CTkTextbox(
            log_frame, height=200, wrap="word",
            font=ctk.CTkFont(family="Consolas", size=11))
        self.log_text.pack(fill="both", expand=True, padx=4, pady=(0, 4))

        # -- Single-pair preview pane
        sp_frame = ctk.CTkFrame(bottom)
        bottom.add(sp_frame, weight=2)
        ctk.CTkLabel(sp_frame, text="Single-pair cross-correlation curve",
                     font=ctk.CTkFont(weight="bold")).pack(
            anchor="w", padx=8, pady=(6, 0))

        ctrl = ctk.CTkFrame(sp_frame, fg_color="transparent")
        ctrl.pack(fill="x", padx=6, pady=(0, 4))
        ctk.CTkLabel(ctrl, text="ROI A:").pack(side="left")
        self.sp_roiA_var = tk.StringVar(value="0")
        ctk.CTkEntry(ctrl, textvariable=self.sp_roiA_var, width=70).pack(
            side="left", padx=(2, 8))
        ctk.CTkLabel(ctrl, text="ROI B:").pack(side="left")
        self.sp_roiB_var = tk.StringVar(value="1")
        ctk.CTkEntry(ctrl, textvariable=self.sp_roiB_var, width=70).pack(
            side="left", padx=(2, 8))
        ctk.CTkLabel(ctrl, text="lag window (s):").pack(side="left")
        self.sp_lag_var = tk.StringVar(value=str(DEFAULT_MAX_LAG_S))
        ctk.CTkEntry(ctrl, textvariable=self.sp_lag_var, width=60).pack(
            side="left", padx=(2, 8))
        self.sp_event_var = tk.StringVar(value="full recording")
        ctk.CTkLabel(ctrl, text="event:").pack(side="left")
        self.sp_event_combo = ctk.CTkComboBox(
            ctrl, variable=self.sp_event_var, state="readonly",
            width=200, values=["full recording"])
        self.sp_event_combo.pack(side="left", padx=(2, 8))
        self.sp_plot_btn = ctk.CTkButton(
            ctrl, text="Plot", width=80, command=self._on_plot_single_pair,
            state="disabled")
        self.sp_plot_btn.pack(side="left")

        self.sp_fig = plt.Figure(figsize=(5.5, 3.0), tight_layout=True)
        self.sp_ax = self.sp_fig.add_subplot(111)
        self._sp_placeholder("Pick two ROIs and click Plot.")
        self.sp_canvas = FigureCanvasTkAgg(self.sp_fig, master=sp_frame)
        tb_frame = ctk.CTkFrame(sp_frame, fg_color="transparent")
        tb_frame.pack(fill="x")
        self.sp_toolbar = NavigationToolbar2Tk(
            self.sp_canvas, tb_frame, pack_toolbar=False)
        self.sp_toolbar.update()
        restyle_matplotlib_toolbar(self.sp_toolbar)
        self.sp_toolbar.pack(side="left", fill="x")
        from ... import plot_data_export as _pde
        ctk.CTkButton(
            tb_frame, text="Save data...", width=110,
            command=lambda: _pde.save_figure_data(
                self.sp_fig, tb_frame, "single_pair_xcorr"),
        ).pack(side="left", padx=(8, 0))
        self.sp_canvas.get_tk_widget().pack(fill="both", expand=True)

    # -- Single-pair preview helpers ---------------------------------------

    def _sp_placeholder(self, text: str) -> None:
        self.sp_ax.clear()
        self.sp_ax.set_axis_off()
        self.sp_ax.text(0.5, 0.5, text, ha="center", va="center",
                        transform=self.sp_ax.transAxes)

    def _get_cached_dff(self):
        """Return (dff_memmap, T, N_kept), reopening only when path/prefix
        changes."""
        if self._plane0 is None:
            raise RuntimeError("No plane0 selected.")
        prefix = self.prefix_var.get().strip() or DEFAULT_PREFIX
        key = (str(self._plane0), prefix)
        if self._dff_cache is None or self._dff_cache_key != key:
            dff, T, N = xc._open_dff_memmap(self._plane0, prefix)
            self._dff_cache = (dff, T, N)
            self._dff_cache_key = key
        return self._dff_cache

    def _on_plot_single_pair(self) -> None:
        try:
            dff, T, N = self._get_cached_dff()
        except Exception as e:
            messagebox.showerror("dF/F unavailable", str(e))
            return
        # Parse ROI indices
        try:
            iA = int(self.sp_roiA_var.get())
            iB = int(self.sp_roiB_var.get())
        except ValueError:
            messagebox.showerror(
                "Bad ROI index", "ROI A and ROI B must be integers.")
            return
        if not (0 <= iA < N and 0 <= iB < N):
            messagebox.showerror(
                "Out of range", f"ROI indices must be in [0, {N - 1}].")
            return
        try:
            max_lag = float(self.sp_lag_var.get())
        except ValueError:
            max_lag = DEFAULT_MAX_LAG_S
        try:
            fps = float(self.fps_var.get())
        except ValueError:
            fps = self._fps

        # Resolve event crop (or full recording).
        ev_label = self.sp_event_var.get()
        f0, f1 = 0, T
        ev_title = "full recording"
        if ev_label != "full recording":
            try:
                ev_idx = int(ev_label.split()[1])
                s_sec, e_sec = self._event_windows[ev_idx]
                f0 = max(0, int(round(s_sec * fps)))
                f1 = min(T, int(round(e_sec * fps)))
                ev_title = (f"event {ev_idx:04d}  [{s_sec:.2f}-{e_sec:.2f}s, "
                            f"{f1 - f0} frames]")
            except (ValueError, IndexError):
                ev_title = "full recording"

        if f1 - f0 < 4:
            messagebox.showerror(
                "Window too short",
                f"Selected window has only {f1 - f0} frames.")
            return
        # Cap max_lag to window length.
        eff_max_lag = min(max_lag, max(0.0, (f1 - f0 - 1) / float(fps)))
        if eff_max_lag <= 0:
            messagebox.showerror(
                "Lag too large",
                "Lag window is wider than the available data window.")
            return

        sigA = np.asarray(dff[f0:f1, iA], dtype=np.float32)
        sigB = np.asarray(dff[f0:f1, iB], dtype=np.float32)
        try:
            lags_sec, r = xc.single_pair_xcorr_curve(
                sigA, sigB, fps, max_lag_seconds=eff_max_lag)
        except Exception as e:
            messagebox.showerror("xcorr failed", str(e))
            return

        peak_idx = int(np.argmax(r))
        peak_lag = float(lags_sec[peak_idx])
        peak_r = float(r[peak_idx])
        zero_idx = int(np.argmin(np.abs(lags_sec)))
        zero_r = float(r[zero_idx])

        ax = self.sp_ax
        ax.clear()
        ax.set_axis_on()
        ax.plot(lags_sec, r, lw=1.2, color="tab:blue")
        ax.axhline(0, color="black", lw=0.5, alpha=0.4)
        ax.axvline(0, color="black", lw=0.5, alpha=0.4)
        ax.axvline(peak_lag, color="tab:red", lw=1.0, ls="--",
                   label=f"peak: lag={peak_lag:.3f}s, r={peak_r:.3f}")
        ax.scatter([0.0], [zero_r], color="tab:green", zorder=5,
                   label=f"lag=0:  r={zero_r:.3f}")
        ax.set_xlabel("lag (s)  [+ = ROI A leads]")
        # NOTE: This is Pearson r on filtered dF/F. Indicator kinetics
        # (GCaMP rise + decay) convolve every spike with a ~tens-to-
        # hundreds-of-ms exponential, so the visible CCG peak is
        # blurred by the indicator's autocorrelation -- the apparent
        # lag resolution is indicator-limited, not frame-rate-limited.
        # Pearson on raw fluorescence also inflates correlation
        # magnitudes via this temporal blur (Yatsenko/Mishne, eLife
        # 2021, doi:10.7554/eLife.68046). To get an indicator-
        # independent CCG, compute on deconvolved/binned spike rates.
        ax.set_ylabel("Pearson r on filtered dF/F\n"
                      "(indicator kinetics blur peak; lag resolution\n"
                      "is indicator- not frame-rate-limited)")
        ax.set_title(f"ROI {iA}  vs  ROI {iB}  -  {ev_title}",
                     fontsize=10)
        ax.legend(fontsize=8, loc="best", frameon=True)
        ax.grid(True, alpha=0.3)
        self.sp_canvas.draw_idle()
        try:
            self.sp_toolbar.update()
        except Exception:
            pass
        self._log(f"[single-pair] ROI {iA} x ROI {iB}  ({ev_title})  "
                  f"peak_lag={peak_lag:.3f}s  peak_r={peak_r:.3f}  "
                  f"zero_r={zero_r:.3f}")

    # -- State plumbing ----------------------------------------------------

    def _on_plane0_broadcast(self, plane0) -> None:
        if plane0 is None:
            return
        self.path_var.set(str(plane0))
        self._set_plane0(Path(plane0))

    def _on_browse(self) -> None:
        path = filedialog.askdirectory(
            title="Select Suite2p plane0 folder",
            initialdir=self.path_var.get() or str(Path.home()))
        if not path:
            return
        self.path_var.set(path)
        self._set_plane0(Path(path))

    def _repoint_after_copy(self, scratch: Path, final: Path) -> None:
        """BatchTab calls this after a row's scratch->HDD copy. If
        ``_dff_cache`` is holding a memmap whose file lives under
        ``scratch``, drop it so Windows releases the handle before the
        bulk rmtree fires. The cache is reopened lazily against the
        HDD twin on the next ``_get_cached_dff`` call.

        Without this hook, the open memmap kept the scratch
        ``r0p7*dff.memmap.float32`` locked across recordings; on the
        last row of a batch the janitor's rmtree never succeeded
        until the user navigated away or closed the app.
        """
        cache = self._dff_cache
        if cache is None:
            return
        dff = cache[0] if isinstance(cache, tuple) else cache
        fname = getattr(dff, "filename", None)
        if fname is None:
            return
        try:
            Path(fname).relative_to(scratch)
        except (TypeError, ValueError, OSError):
            return
        self._dff_cache = None
        self._dff_cache_key = None

    def _set_plane0(self, plane0: Path) -> None:
        self._plane0 = plane0
        # Reset the dF/F memmap cache so the next preview reopens.
        self._dff_cache = None
        self._dff_cache_key = None
        try:
            fps = float(utils.get_fps_from_notes(str(plane0),
                                                 default_fps=15.07))
            self._fps = fps
            self.fps_var.set(f"{fps:.3f}")
        except Exception:
            pass
        self._on_refresh_events()
        ok = self._inputs_ready()
        self.full_btn.configure(state="normal" if ok else "disabled")
        self.per_event_btn.configure(
            state="normal" if (ok and self._event_windows) else "disabled")
        # Single-pair preview only needs the dF/F memmap, not the cluster files.
        prefix = self.prefix_var.get().strip() or DEFAULT_PREFIX
        sp_ok = (plane0 is not None
                 and (plane0 / f"{prefix}dff.memmap.float32").exists())
        self.sp_plot_btn.configure(state="normal" if sp_ok else "disabled")
        # Violin plot needs only the per-pair summary CSVs to exist on disk;
        # we keep the button enabled whenever a plane0 is selected and check
        # the actual files at click time.
        self.violin_btn.configure(
            state="normal" if plane0 is not None else "disabled")
        self.export_figs_btn.configure(
            state="normal" if plane0 is not None else "disabled")
        if ok:
            self.status_var.set(f"Ready. ({plane0})")
        else:
            self.status_var.set("Cluster files or dF/F memmap missing.")

    def _inputs_ready(self) -> bool:
        if self._plane0 is None:
            return False
        prefix = self.prefix_var.get().strip() or DEFAULT_PREFIX
        cfolder = self.cfolder_var.get().strip() or DEFAULT_CLUSTER_FOLDER
        cluster_dir = (self._plane0 / f"{prefix}cluster_results" / cfolder)
        if not (self._plane0 / f"{prefix}dff.memmap.float32").exists():
            return False
        if not cluster_dir.exists():
            return False
        roi_files = [f for f in cluster_dir.glob("*_rois.npy")
                     if "manual_combined" not in f.stem.lower()]
        return len(roi_files) >= 1

    def _on_refresh_events(self) -> None:
        if self._plane0 is None:
            self._apply_event_windows([])
            return
        # Prefer the in-memory publish from Tab 5 if it matches the
        # currently-loaded plane0 -- it's the freshest source. Otherwise
        # fall back to the EventWindows sheet on disk (cold-start path).
        evts = self._event_windows_from_state()
        if evts is None:
            evts = _read_event_windows_from_summary(self._plane0)
        self._apply_event_windows(evts)

    def _event_windows_from_state(self):
        """Return Tab 5's in-memory ``[(start_s, end_s), ...]`` for the
        current plane0, or None if no matching publish is cached."""
        results = getattr(self.state, "event_results", None) if self.state \
            else None
        if not results:
            return None
        if Path(results.get("plane0", "")) != Path(self._plane0):
            return None
        ev = results.get("event_windows")
        if ev is None or len(ev) == 0:
            return []
        return [(float(t0), float(t1)) for t0, t1 in ev]

    def _apply_event_windows(self, evts) -> None:
        """Update ``_event_windows`` and the dependent UI bits."""
        evts = list(evts) if evts else []
        self._event_windows = evts
        if not evts and self._plane0 is None:
            self.event_count_var.set("events: -")
        else:
            self.event_count_var.set(f"events: {len(evts)}")
        self.per_event_btn.configure(
            state="normal" if (self._inputs_ready() and evts) else "disabled")
        labels = ["full recording"] + [
            f"event {i:04d}  [{s:.2f}-{e:.2f}s]"
            for i, (s, e) in enumerate(evts)
        ]
        self.sp_event_combo.configure(values=labels)
        if self.sp_event_var.get() not in labels:
            self.sp_event_var.set("full recording")

    def _on_event_results_broadcast(self, results: dict) -> None:
        """Tab 5 just published a fresh event-detection result. If it's
        for the plane0 we're showing, refresh in-memory."""
        if not results or self._plane0 is None:
            return
        if Path(results.get("plane0", "")) != Path(self._plane0):
            return
        ev = results.get("event_windows")
        evts = ([(float(t0), float(t1)) for t0, t1 in ev]
                if ev is not None else [])
        self._apply_event_windows(evts)

    def _on_export_figures(self, quiet: bool = False) -> None:
        """Re-render Tab 7's violin pair plots into
        ``<save_folder>/calliope_figures/crosscorrelation/`` as PNG +
        SVG, plus update the manifest. Pulls per-pair data from the
        on-disk ``cross_correlation_full`` (and optionally
        ``cross_correlation_per_event``) outputs, so this is fast --
        no recompute of the pairwise sweep.

        ``quiet=True`` (Tab 0 batch toggle) suppresses messageboxes.
        """
        if self._plane0 is None:
            if not quiet:
                messagebox.showinfo(
                    "No plane0", "Pick a plane0 folder first.")
            return
        prefix = self.prefix_var.get().strip() or DEFAULT_PREFIX
        cfolder = self.cfolder_var.get().strip() or DEFAULT_CLUSTER_FOLDER
        cluster_root = (self._plane0 / f"{prefix}cluster_results" / cfolder)
        full_dir = cluster_root / DEFAULT_FULL_OUTPUT_SUBDIR
        per_event_dir = cluster_root / "cross_correlation_per_event"
        if not full_dir.is_dir():
            if not quiet:
                messagebox.showerror(
                    "No outputs",
                    f"Run the full-recording cross-correlation first.\n\n"
                    f"Expected:\n{full_dir}")
            return
        save_folder = self._plane0.parents[3]
        figures_root = save_folder / "calliope_figures"
        xc_dir = figures_root / "crosscorrelation"
        try:
            xc_dir.mkdir(parents=True, exist_ok=True)
            from ...core.crosscorrelation import _save_violin_pair
            from ...core.export_manifest import write_export_manifest
            from ...core.utils import safe_recording_id
            rec_id = safe_recording_id(self._plane0)
            n_files = 0
            full_pair = _load_pair_data(full_dir)
            _save_violin_pair(
                full_pair, save_path=xc_dir / "full.png",
                title_suffix="full recording")
            n_files += 2  # png + svg
            if per_event_dir.is_dir():
                pe_pair = _load_pair_data(per_event_dir)
                _save_violin_pair(
                    pe_pair, save_path=xc_dir / "per_event.png",
                    title_suffix="per event")
                n_files += 2
            manifest_path = write_export_manifest(
                figures_root,
                rec_id=rec_id, params={
                    "prefix": prefix,
                    "cluster_folder": cfolder,
                },
                plane0=self._plane0, ckpt_path=None,
            )
            self.status_var.set(
                f"Exported xcorr figures -> {xc_dir}")
            if not quiet:
                messagebox.showinfo(
                    "Export complete",
                    f"Wrote {n_files} violin-plot files (PNG + SVG) "
                    f"to:\n  {xc_dir}\n\n"
                    f"Manifest:\n  {manifest_path}")
        except Exception as e:
            self.status_var.set(f"Export failed: {e}")
            if not quiet:
                messagebox.showerror("Export failed", str(e))

    def _on_violin(self) -> None:
        """Open a violin-plot window over the per-pair summary CSVs from the
        most recent full-recording cross-correlation run."""
        if self._plane0 is None:
            messagebox.showinfo("No plane0", "Pick a plane0 folder first.")
            return
        prefix = self.prefix_var.get().strip() or DEFAULT_PREFIX
        cfolder = self.cfolder_var.get().strip() or DEFAULT_CLUSTER_FOLDER
        xcorr_root = (self._plane0 / f"{prefix}cluster_results"
                      / cfolder / DEFAULT_FULL_OUTPUT_SUBDIR)
        if not xcorr_root.is_dir():
            messagebox.showerror(
                "No outputs",
                f"Run the full-recording cross-correlation first.\n\n"
                f"Expected:\n{xcorr_root}")
            return
        try:
            # Resolve recording id robustly (handles both old
            # ``<rec>/suite2p/plane0`` and new
            # ``<rec>/detection/final/suite2p/plane0`` layouts).
            from ...core.utils import infer_recording_id
            recording_label = infer_recording_id(self._plane0)
        except Exception:
            recording_label = ""
        try:
            ViolinWindow(self.winfo_toplevel(), xcorr_root=xcorr_root,
                         recording_label=recording_label)
        except Exception as e:
            messagebox.showerror("Violin plot failed",
                                 f"{e}\n{traceback.format_exc()}")

    def _on_reload_inputs(self) -> None:
        """Drop the cached dF/F memmap and re-check cluster/dF/F files on
        disk. Useful after re-running detection or regenerating cluster
        outputs while the GUI is open."""
        path_str = self.path_var.get().strip()
        plane0 = self._plane0 if self._plane0 is not None else (
            Path(path_str) if path_str else None)
        if plane0 is None:
            messagebox.showinfo(
                "No plane0",
                "Pick a plane0 folder before reloading.")
            return
        # Force the next preview/run to reopen the memmap.
        self._dff_cache = None
        self._dff_cache_key = None
        # Re-resolve fps from notes, re-check inputs, and refresh events.
        self._set_plane0(Path(plane0))
        self._log(f"[reload] dF/F + clusters refreshed for {plane0}")

    # -- Logging -----------------------------------------------------------

    def _log(self, msg: str) -> None:
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")

    def _set_progress(self, done: int, total: int) -> None:
        # CTkProgressBar is always 0..1 (no separate ``maximum`` arg).
        if total <= 0:
            self.progress.set(0)
            return
        self.progress.set(max(0.0, min(1.0, done / total)))

    # -- Run: full recording -----------------------------------------------

    def _gather_params(self):
        prefix = self.prefix_var.get().strip() or DEFAULT_PREFIX
        cfolder = self.cfolder_var.get().strip() or DEFAULT_CLUSTER_FOLDER
        try:
            fps = float(self.fps_var.get())
        except ValueError:
            fps = self._fps
        try:
            max_lag = float(self.maxlag_var.get())
        except ValueError:
            max_lag = DEFAULT_MAX_LAG_S
        try:
            n_shuffles = max(0, int(self.n_shuffles_var.get()))
        except (TypeError, ValueError):
            n_shuffles = 500
        return {
            "prefix": prefix, "cluster_folder": cfolder, "fps": fps,
            "max_lag_seconds": max_lag,
            "zero_lag": bool(self.zero_lag_var.get()),
            "use_gpu": bool(self.gpu_var.get()),
            "n_shuffles": n_shuffles,
            "compute_partial": bool(self.partial_corr_var.get()),
        }

    def _disable_run(self) -> None:
        self.full_btn.configure(state="disabled")
        self.per_event_btn.configure(state="disabled")
        # Worker is starting -- arm the abort button.
        self._abort_event.clear()
        self.abort_btn.configure(state="normal")

    def _enable_run(self) -> None:
        ok = self._inputs_ready()
        self.full_btn.configure(state="normal" if ok else "disabled")
        self.per_event_btn.configure(
            state="normal" if (ok and self._event_windows) else "disabled")
        # No worker running -- abort makes no sense.
        self.abort_btn.configure(state="disabled")
        self._abort_event.clear()

    def _on_abort(self) -> None:
        """User clicked Abort. Set the flag; the next progress_cb call
        from inside the cross-correlation routine raises RunAborted."""
        if self._worker is None or not self._worker.is_alive():
            return
        self._abort_event.set()
        self.abort_btn.configure(state="disabled")
        self.status_var.set("Aborting... waiting for current batch to finish.")
        self._log("[abort] requested")

    def _on_run_full(self) -> None:
        if self._plane0 is None or not self._inputs_ready():
            messagebox.showerror("Not ready", "Inputs incomplete.")
            return
        if self._worker is not None and self._worker.is_alive():
            messagebox.showinfo("Busy", "A run is already in progress.")
            return
        params = self._gather_params()
        plane0 = self._plane0
        self._disable_run()
        self.progress.set(0)
        self.status_var.set("Running full-recording cross-correlation...")
        self._log(f"\n[full] plane0={plane0}")
        self._log(f"[full] params={params}")

        def progress_cb(done, total, label):
            if self._abort_event.is_set():
                raise RunAborted()
            self._q.put(("progress", (done, total, label)))

        def worker():
            try:
                out = xc.run_cluster_xcorr_full_fast(
                    plane0, progress_cb=progress_cb, **params,
                )
                self._q.put(("done_full", str(out)))
            except RunAborted:
                self._q.put(("aborted", "full"))
            except Exception as e:
                self._q.put(("error", f"{e}\n{traceback.format_exc()}"))

        self._worker = threading.Thread(target=worker, daemon=True)
        self._worker.start()

    def _on_run_per_event(self) -> None:
        if self._plane0 is None or not self._inputs_ready():
            messagebox.showerror("Not ready", "Inputs incomplete.")
            return
        if not self._event_windows:
            messagebox.showinfo(
                "No events",
                "No event windows found in the recording summary "
                "(*summary*.xlsx). Run event detection first.")
            return
        if self._worker is not None and self._worker.is_alive():
            messagebox.showinfo("Busy", "A run is already in progress.")
            return
        params = self._gather_params()
        # Per-event runner accepts n_shuffles for API symmetry but the
        # current implementation does not compute the shuffle null per
        # event (would require per-window jitter). Strip it to make the
        # no-op explicit and avoid a misleading n_shuffles= in the log.
        n_shuf = params.pop("n_shuffles", 0)
        if n_shuf:
            self._log(f"[per-event] note: n_shuffles={n_shuf} ignored "
                      "(per-event circular-shift null not implemented).")
        # Partial correlation is only meaningful on the full recording
        # (per-event windows are usually too short for a conditioned
        # precision matrix). Strip the flag so the per-event call site
        # doesn't see an unknown kwarg.
        partial_flag = params.pop("compute_partial", False)
        if partial_flag:
            self._log("[per-event] note: partial correlation only supported "
                      "for full-recording runs; ignoring for per-event.")
        plane0 = self._plane0
        evts = list(self._event_windows)
        self._disable_run()
        self._set_progress(0, len(evts))
        self.status_var.set(
            f"Running per-event cross-correlation ({len(evts)} events)...")
        self._log(f"\n[per-event] plane0={plane0}")
        self._log(f"[per-event] params={params}  events={len(evts)}")

        def progress_cb(done, total, label):
            if self._abort_event.is_set():
                raise RunAborted()
            self._q.put(("progress", (done, total, label)))

        def worker():
            try:
                out = xc.run_cluster_xcorr_per_event_fast(
                    plane0, event_windows=evts,
                    progress_cb=progress_cb, **params,
                )
                self._q.put(("done_per_event", str(out)))
            except RunAborted:
                self._q.put(("aborted", "per_event"))
            except Exception as e:
                self._q.put(("error", f"{e}\n{traceback.format_exc()}"))

        self._worker = threading.Thread(target=worker, daemon=True)
        self._worker.start()

    # -- Queue handlers ----------------------------------------------------

    def _on_xc_progress(self, payload) -> None:
        done, total, label = payload
        self._set_progress(done, total)
        self.status_var.set(f"{label} ({done}/{total})")

    def _on_xc_done_full(self, payload: str) -> None:
        self._set_progress(1, 1)
        self.status_var.set(f"Done. Output: {payload}")
        self._log(f"[full] done -> {payload}")
        self._enable_run()
        # Tab 0's batch runner subscribes to advance to spatial.
        try:
            self.state.set_xcorr_ready(Path(payload))
        except Exception as e:
            print(f"[GUI] xcorr_ready publish failed: {e}")

    def _on_xc_done_per_event(self, payload: str) -> None:
        # CTkProgressBar uses 0..1, so finish by pinning to 1.0.
        self._set_progress(1, 1)
        self.status_var.set(f"Done. Output: {payload}")
        self._log(f"[per-event] done -> {payload}")
        self._enable_run()
        try:
            self.state.set_xcorr_ready(Path(payload))
        except Exception as e:
            print(f"[GUI] xcorr_ready publish failed: {e}")

    def _on_xc_aborted(self, payload: str) -> None:
        self.status_var.set("Aborted by user.")
        self._log(f"[abort] {payload} run aborted by user")
        self._enable_run()

    def _on_xc_error(self, payload: str) -> None:
        self.status_var.set("Cross-correlation failed.")
        self._log("[error] " + payload)
        self._enable_run()
        messagebox.showerror(
            "Cross-correlation failed",
            payload.split("\n", 1)[0])


def main() -> None:
    """Stand-alone launcher for hot-running Tab 7 without the full
    ``pipeline_gui`` app. Mirrors the dark-mode + ttk-restyle setup
    ``pipeline_gui.main`` does so the tab still reads correctly.
    """
    from ...gui_common import apply_ttk_dark_theme
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    root = ctk.CTk()
    apply_ttk_dark_theme(root)
    root.title("CalLIOPE - Cross-correlation")
    root.geometry("1100x720")
    tab = CrossCorrelationTab(root, state=None)
    tab.pack(fill="both", expand=True)
    root.mainloop()


if __name__ == "__main__":
    main()
